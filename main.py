import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# dictionaries
products_per_supplier = {}
total_value_per_supplier = {}
# dictionary for product under 20
products_under_20_inv = {}

# this will print total rows
# print(product_list.max_row)

# this line will start from line 2 as first line is header and we do not want this data
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_of_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_of_products + 1
    else:
        # print("adding new supplier")
        products_per_supplier[supplier_name] = 1
    # print(products_per_supplier)

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# print(products_per_supplier)
# print(total_value_per_supplier)

    # logic for products inventory under 10
    if inventory < 20:
        products_under_20_inv[product_num] = inventory

# print(products_under_20_inv)

    #add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_20_inv)

inv_file.save("inventory_with_total_value.xlsx")